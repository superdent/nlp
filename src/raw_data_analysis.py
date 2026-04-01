"""
Explorative Datenanalyse (EDA) für Sentimentanalyse-Projekt
Erzeugt Excel + Markdown-Report mit PNG-Grafiken
Liest alle Records (kein Limit), arbeitet speicherschonend ohne DataFrame
"""

import json
import gc
from datetime import datetime
from pathlib import Path
from collections import Counter
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

PROJECT_ROOT = Path(".")

INPUT_FILES = {
    "Movies & TV": "data/raw/Movies_and_TV.jsonl",
    "All Beauty": "data/raw/All_Beauty.jsonl",
    "Office Products": "data/raw/Office_Products.jsonl"
}

SPLIT_RATIOS = (0.7, 0.15, 0.15)

TEXTLEN_BUCKETS = [
    (0, 0, "0 (leer)"),
    (1, 10, "1-10"),
    (11, 50, "11-50"),
    (51, 100, "51-100"),
    (101, 500, "101-500"),
    (501, 1000, "501-1000"),
    (1001, 5000, "1001-5000"),
    (5001, float('inf'), "5001+"),
]


def bucket_index(length):
    for i, (lo, hi, _) in enumerate(TEXTLEN_BUCKETS):
        if lo <= length <= hi:
            return i
    return len(TEXTLEN_BUCKETS) - 1


def analyze_category(cat_name, filepath):
    full_path = PROJECT_ROOT / filepath
    print(f"\n{'='*60}")
    print(f"Analysiere: {cat_name}")
    print(f"{'='*60}")

    total_records = 0
    invalid_ratings = 0
    json_errors = 0
    null_title = 0
    null_text = 0
    null_rating = 0
    empty_texts = 0

    valid_ratings = []
    rating_counts = Counter()
    textlen_sum = 0
    wordcount_sum = 0
    textlen_min = float('inf')
    textlen_max = 0
    wordcount_min = float('inf')
    wordcount_max = 0
    bucket_counts = [0] * len(TEXTLEN_BUCKETS)

    with open(full_path, "r", encoding="utf-8") as f:
        for line in f:
            total_records += 1
            if total_records % 5_000_000 == 0:
                print(f"  {total_records:,} Records gelesen... {datetime.now().strftime('%H:%M:%S')}")

            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                json_errors += 1
                continue

            title = record.get('title')
            text = record.get('text')
            rating = record.get('rating')

            if title is None:
                null_title += 1
            if text is None:
                null_text += 1
            if rating is None:
                null_rating += 1

            if rating not in [1, 2, 3, 4, 5]:
                invalid_ratings += 1
                continue

            text_str = text or ''
            tlen = len(text_str)
            wcount = len(text_str.split())

            if tlen == 0:
                empty_texts += 1

            rating_counts[int(rating)] += 1
            textlen_sum += tlen
            wordcount_sum += wcount
            if tlen < textlen_min:
                textlen_min = tlen
            if tlen > textlen_max:
                textlen_max = tlen
            if wcount < wordcount_min:
                wordcount_min = wcount
            if wcount > wordcount_max:
                wordcount_max = wcount
            bucket_counts[bucket_index(tlen)] += 1

            valid_ratings.append(int(rating))

    valid_count = len(valid_ratings)
    print(f"  Gesamt: {total_records:,}, Valide: {valid_count:,}, Ungueltig: {invalid_ratings:,}")

    # Split-Verteilung
    n = valid_count
    train_size = int(n * SPLIT_RATIOS[0])
    val_size = int(n * SPLIT_RATIOS[1])

    splits = {
        'Train': valid_ratings[:train_size],
        'Val': valid_ratings[train_size:train_size + val_size],
        'Test': valid_ratings[train_size + val_size:],
    }

    split_stats = {}
    for split_name, ratings in splits.items():
        sc = Counter(ratings)
        split_stats[split_name] = {
            'total': len(ratings),
            'r1': sc.get(1, 0),
            'r2': sc.get(2, 0),
            'r3': sc.get(3, 0),
            'r4': sc.get(4, 0),
            'r5': sc.get(5, 0),
        }

    del valid_ratings
    gc.collect()

    return {
        'category': cat_name,
        'total_records': total_records,
        'json_errors': json_errors,
        'invalid_ratings': invalid_ratings,
        'valid_records': valid_count,
        'null_title': null_title,
        'null_text': null_text,
        'null_rating': null_rating,
        'empty_texts': empty_texts,
        'r1': rating_counts.get(1, 0),
        'r2': rating_counts.get(2, 0),
        'r3': rating_counts.get(3, 0),
        'r4': rating_counts.get(4, 0),
        'r5': rating_counts.get(5, 0),
        'textlen_sum': textlen_sum,
        'textlen_min': textlen_min if textlen_min != float('inf') else 0,
        'textlen_max': textlen_max,
        'wordcount_sum': wordcount_sum,
        'wordcount_min': wordcount_min if wordcount_min != float('inf') else 0,
        'wordcount_max': wordcount_max,
        'bucket_counts': bucket_counts,
        'split_stats': split_stats,
    }


def create_plots(stats, figures_dir):
    cat_name = stats['category']

    # Bewertungsverteilung
    fig, ax = plt.subplots(figsize=(8, 5))
    ratings = [1, 2, 3, 4, 5]
    counts = [stats[f'r{r}'] for r in ratings]
    ax.bar(ratings, counts, color='steelblue', edgecolor='black')
    ax.set_xlabel('Bewertung', fontsize=11)
    ax.set_ylabel('Anzahl', fontsize=11)
    ax.set_title(f'Verteilung der Bewertungen: {cat_name}', fontsize=12, fontweight='bold')
    ax.set_xticks(ratings)
    ax.grid(axis='y', alpha=0.3)

    rating_file = f"{cat_name.replace(' ', '_')}_rating.png"
    fig.savefig(figures_dir / rating_file, dpi=100, bbox_inches='tight')
    plt.close(fig)

    # Textlaengen-Histogramm (aus Buckets)
    fig, ax = plt.subplots(figsize=(8, 5))
    labels = [b[2] for b in TEXTLEN_BUCKETS]
    ax.bar(range(len(labels)), stats['bucket_counts'], color='coral', edgecolor='black', alpha=0.7)
    ax.set_xlabel('Textlaenge (Zeichen)', fontsize=11)
    ax.set_ylabel('Anzahl', fontsize=11)
    ax.set_title(f'Verteilung der Textlaengen: {cat_name}', fontsize=12, fontweight='bold')
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=45, ha='right')
    ax.grid(axis='y', alpha=0.3)

    textlen_file = f"{cat_name.replace(' ', '_')}_textlen.png"
    fig.savefig(figures_dir / textlen_file, dpi=100, bbox_inches='tight')
    plt.close(fig)

    return rating_file, textlen_file


def generate_markdown_report(all_stats, all_plot_files, figures_dirname):
    md = []
    md.append("# EDA-Report: Sentimentanalyse")
    md.append("")
    md.append("Amazon-Reviews aus 4 Produktkategorien (alle Records, kein Limit)")
    md.append("")

    # Uebersichtstabelle
    md.append("## Uebersichtstabelle")
    md.append("")
    md.append("| Kategorie | Records gesamt | Valide | Ungueltige Ratings | Leere Texte | Ø Textlaenge | Ø Woerter |")
    md.append("|-----------|---------------|--------|-------------------|-------------|-------------|----------|")
    for s in all_stats:
        avg_tlen = s['textlen_sum'] / s['valid_records'] if s['valid_records'] > 0 else 0
        avg_wc = s['wordcount_sum'] / s['valid_records'] if s['valid_records'] > 0 else 0
        md.append(f"| {s['category']} | {s['total_records']:,} | {s['valid_records']:,} | {s['invalid_ratings']:,} | {s['empty_texts']:,} | {avg_tlen:.0f} | {avg_wc:.0f} |")
    md.append("")
    md.append("---")
    md.append("")

    # Pro Kategorie
    for s in all_stats:
        cat_name = s['category']
        rating_file, textlen_file = all_plot_files[cat_name]
        avg_tlen = s['textlen_sum'] / s['valid_records'] if s['valid_records'] > 0 else 0
        avg_wc = s['wordcount_sum'] / s['valid_records'] if s['valid_records'] > 0 else 0

        md.append(f"## {cat_name}")
        md.append("")
        md.append("### Statistiken")
        md.append(f"- Records gesamt: {s['total_records']:,}")
        md.append(f"- Valide Records: {s['valid_records']:,}")
        md.append(f"- Ungueltige Ratings: {s['invalid_ratings']:,}")
        md.append(f"- JSON-Fehler: {s['json_errors']:,}")
        md.append(f"- Fehlende Werte: title={s['null_title']}, text={s['null_text']}, rating={s['null_rating']}")
        md.append(f"- Leere Texte: {s['empty_texts']:,}")
        md.append("")

        md.append("### Bewertungsverteilung")
        for i in range(1, 6):
            count = s[f'r{i}']
            pct = count / s['valid_records'] * 100 if s['valid_records'] > 0 else 0
            md.append(f"- {i} Stern: {count:,} ({pct:.1f}%)")
        md.append("")

        md.append("### Textlaenge")
        md.append(f"- Mittelwert: {avg_tlen:.0f} Zeichen / {avg_wc:.0f} Woerter")
        md.append(f"- Min: {s['textlen_min']} / Max: {s['textlen_max']}")
        md.append("")

        md.append("### Textlaengen-Buckets")
        md.append("| Bucket | Anzahl |")
        md.append("|--------|--------|")
        for (_, _, label), count in zip(TEXTLEN_BUCKETS, s['bucket_counts']):
            md.append(f"| {label} | {count:,} |")
        md.append("")

        md.append(f"### Bewertungsverteilung")
        md.append(f"![Bewertungsverteilung]({figures_dirname}/{rating_file})")
        md.append("")
        md.append(f"### Textlaengenverteilung")
        md.append(f"![Textlaengen]({figures_dirname}/{textlen_file})")
        md.append("")
        md.append("---")
        md.append("")

    # Split-Verteilung
    md.append("## Split-Verteilung (70/15/15)")
    md.append("")
    md.append("| Kategorie | Split | Gesamt | 1 Stern | 2 Sterne | 3 Sterne | 4 Sterne | 5 Sterne |")
    md.append("|-----------|-------|--------|---------|----------|----------|----------|----------|")
    for s in all_stats:
        for split_name in ['Train', 'Val', 'Test']:
            ss = s['split_stats'][split_name]
            md.append(f"| {s['category']} | {split_name} | {ss['total']:,} | {ss['r1']:,} | {ss['r2']:,} | {ss['r3']:,} | {ss['r4']:,} | {ss['r5']:,} |")
    md.append("")

    return "\n".join(md)


def get_next_filename(directory, prefix, extension):
    directory = Path(directory)
    directory.mkdir(parents=True, exist_ok=True)
    existing = list(directory.glob(f"{prefix}_*.{extension}"))
    if not existing:
        return directory / f"{prefix}_01.{extension}"
    nums = []
    for f in existing:
        try:
            num = int(f.stem.split('_')[-1])
            nums.append(num)
        except ValueError:
            pass
    next_num = max(nums) + 1 if nums else 1
    return directory / f"{prefix}_{next_num:02d}.{extension}"


def write_excel(all_stats):
    output_path = get_next_filename(
        PROJECT_ROOT / "documentation", "raw_data_analysis", "xlsx"
    )

    wb = Workbook()
    hf = Font(bold=True, size=11, name='Arial')
    df = Font(size=11, name='Arial')
    la = Alignment(horizontal='left')

    # Sheet 1: Uebersicht
    ws1 = wb.active
    ws1.title = "Uebersicht"
    headers = [
        "Kategorie", "Records gesamt", "JSON-Fehler", "Ungueltige Ratings", "Valide Records",
        "Fehlend: Title", "Fehlend: Text", "Fehlend: Rating", "Leere Texte",
        "1 Stern", "2 Sterne", "3 Sterne", "4 Sterne", "5 Sterne",
        "Textlaenge Summe", "Textlaenge Min", "Textlaenge Max",
        "Wortanzahl Summe", "Wortanzahl Min", "Wortanzahl Max",
    ]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = hf
        c.alignment = la
    for row_idx, s in enumerate(all_stats, 2):
        values = [
            s['category'], s['total_records'], s['json_errors'], s['invalid_ratings'], s['valid_records'],
            s['null_title'], s['null_text'], s['null_rating'], s['empty_texts'],
            s['r1'], s['r2'], s['r3'], s['r4'], s['r5'],
            s['textlen_sum'], s['textlen_min'], s['textlen_max'],
            s['wordcount_sum'], s['wordcount_min'], s['wordcount_max'],
        ]
        for col, v in enumerate(values, 1):
            c = ws1.cell(row=row_idx, column=col, value=v)
            c.font = df
            c.alignment = la
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = max(
            len(str(ws1.cell(row=1, column=col).value)) + 4, 14
        )

    # Sheet 2: Textlaengen-Histogramm
    ws2 = wb.create_sheet("Textlaengen-Histogramm")
    hist_headers = ["Kategorie"] + [b[2] for b in TEXTLEN_BUCKETS]
    for col, h in enumerate(hist_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hf
        c.alignment = la
    for row_idx, s in enumerate(all_stats, 2):
        ws2.cell(row=row_idx, column=1, value=s['category']).font = df
        for col, count in enumerate(s['bucket_counts'], 2):
            c = ws2.cell(row=row_idx, column=col, value=count)
            c.font = df
            c.alignment = la
    for col in range(1, len(hist_headers) + 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = max(
            len(str(ws2.cell(row=1, column=col).value)) + 4, 14
        )

    # Sheet 3: Split-Verteilung
    ws3 = wb.create_sheet("Split-Verteilung")
    split_headers = ["Kategorie", "Split", "Gesamt", "1 Stern", "2 Sterne", "3 Sterne", "4 Sterne", "5 Sterne"]
    for col, h in enumerate(split_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = hf
        c.alignment = la
    row_idx = 2
    for s in all_stats:
        for split_name in ['Train', 'Val', 'Test']:
            ss = s['split_stats'][split_name]
            values = [s['category'], split_name, ss['total'], ss['r1'], ss['r2'], ss['r3'], ss['r4'], ss['r5']]
            for col, v in enumerate(values, 1):
                c = ws3.cell(row=row_idx, column=col, value=v)
                c.font = df
                c.alignment = la
            row_idx += 1
    for col in range(1, len(split_headers) + 1):
        ws3.column_dimensions[ws3.cell(row=1, column=col).column_letter].width = max(
            len(str(ws3.cell(row=1, column=col).value)) + 4, 14
        )

    wb.save(output_path)
    print(f"\nExcel gespeichert: {output_path}")


def main():
    doc_dir = PROJECT_ROOT / "documentation"
    figures_dirname = "eda_figures"
    figures_dir = doc_dir / figures_dirname
    figures_dir.mkdir(parents=True, exist_ok=True)

    print("\n" + "="*60)
    print("EXPLORATIVE DATENANALYSE (alle Records)")
    print("="*60 + "\n")

    all_stats = []
    all_plot_files = {}

    for cat_name, filepath in INPUT_FILES.items():
        stats = analyze_category(cat_name, filepath)
        all_stats.append(stats)

        rating_file, textlen_file = create_plots(stats, figures_dir)
        all_plot_files[cat_name] = (rating_file, textlen_file)

    print(f"\n{'='*60}")
    print("ZUSAMMENFASSUNG")
    print(f"{'='*60}")
    for s in all_stats:
        print(f"  {s['category']}: {s['valid_records']:,} valide von {s['total_records']:,} gesamt")

    write_excel(all_stats)

    report_md = generate_markdown_report(all_stats, all_plot_files, figures_dirname)
    report_path = doc_dir / "2_eda_result.md"
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report_md)
    print(f"Markdown-Report gespeichert: {report_path}")


if __name__ == "__main__":
    main()