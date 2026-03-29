"""
create_appendix_a.py
Erstellt Anhang A: Tabelle mit allen Trainingsläufen.
Liest alle overview.csv Dateien ein und extrahiert relevante Spalten.
"""

from pathlib import Path
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(".")
RESULTS_DIR = PROJECT_ROOT / "results"

MODELS = {
    "Naive Bayes":         PROJECT_ROOT / "results" / "bayes" / "runs" / "overview.csv",
    "SVM":                 PROJECT_ROOT / "results" / "svm" / "runs" / "overview.csv",
    "Logistic Regression": PROJECT_ROOT / "results" / "logistic" / "runs" / "overview.csv",
    "Neural Network":      PROJECT_ROOT / "results" / "neural" / "runs" / "overview.csv",
}

def get_output_path():
    existing_files = list(RESULTS_DIR.glob("appendix_a_*.xlsx"))

    if not existing_files:
        return RESULTS_DIR / "appendix_a_01.xlsx"

    numbers = []
    for f in existing_files:
        try:
            num = int(f.stem.split('_')[-1])
            numbers.append(num)
        except (ValueError, IndexError):
            continue

    next_num = max(numbers) + 1 if numbers else 1
    return RESULTS_DIR / f"appendix_a_{next_num:02d}.xlsx"

def extract_key_hyperparameters(hyperparams_str):
    """Extrahiert wichtigste Hyperparameter aus dem Hyperparameter-String."""
    if not isinstance(hyperparams_str, str):
        return ""

    params = {}
    pairs = hyperparams_str.split(" | ")

    for pair in pairs:
        if ":" in pair:
            key, value = pair.split(":", 1)
            key = key.strip()
            value = value.strip()

            # Nur wichtige Parameter behalten
            if key in ["C", "alpha", "solver", "hidden_units", "dropout_rate", "learning_rate"]:
                params[key] = value

    # Sortiert zusammensetzen
    result = " | ".join([f"{k}: {v}" for k, v in sorted(params.items())])
    return result if result else "default"

rows = []
for model_name, path in MODELS.items():
    if path.exists():
        df = pd.read_csv(path, sep=';')

        # Spalten vorbereiten
        df["Model"] = model_name
        df["Total_Samples"] = df["Train_Samples"] + df["Test_Samples"]
        df["Key_Hyperparameters"] = df["Hyperparameters"].apply(extract_key_hyperparameters)

        # Umbenennung und Spalten wählen
        df_subset = df[["Model", "Total_Samples", "Key_Hyperparameters", "Accuracy",
                        "Precision_Macro", "Recall_Macro", "F1_Macro"]].copy()
        df_subset.columns = ["Model", "Total_Samples", "Key_Hyperparameters",
                             "Accuracy", "Precision", "Recall", "F1"]

        # Auf 4 Dezimalstellen runden
        for col in ["Accuracy", "Precision", "Recall", "F1"]:
            df_subset[col] = df_subset[col].round(4)

        rows.append(df_subset)
        print(f"✓ {model_name}: {path}")
    else:
        print(f"✗ {model_name}: {path} nicht gefunden")

if not rows:
    print("Keine Dateien gefunden. Abbruch.")
else:
    OUTPUT_FILE = get_output_path()
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    df_combined = pd.concat(rows, ignore_index=True)
    df_combined.to_excel(OUTPUT_FILE, index=False)

    # Formatierung
    from openpyxl import load_workbook
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(OUTPUT_FILE)
    print(f"\n✓ Anhang A gespeichert: {OUTPUT_FILE.resolve()}")