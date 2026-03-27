"""
export_results.py
Liest alle Model-Ergebnisse aus overview.csv Dateien ein und schreibt sie in eine gemeinsame Excel-Datei.
Ablegen im Root des Projekts und von dort ausführen.
"""

from pathlib import Path
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(".")
RESULTS_DIR = PROJECT_ROOT / "results"

MODELS = {
    "Naive Bayes":         PROJECT_ROOT / "results" / "bayes" / "runs" / "overview.csv",
    "SVM":                 PROJECT_ROOT / "results" / "svm" / "runs" / "overview.csv",
    "Logistic Regression": PROJECT_ROOT / "results" / "logistic" / "runs" / "overview.csv",
    "Neural Network":      PROJECT_ROOT / "results" / "neural" / "runs" / "overview.csv",
}

def get_output_path():
    existing_files = list(RESULTS_DIR.glob("model_comparison_*.xlsx"))

    if not existing_files:
        return RESULTS_DIR / "model_comparison_01.xlsx"

    numbers = []
    for f in existing_files:
        try:
            num = int(f.stem.split('_')[-1])
            numbers.append(num)
        except (ValueError, IndexError):
            continue

    next_num = max(numbers) + 1 if numbers else 1
    return RESULTS_DIR / f"model_comparison_{next_num:02d}.xlsx"

rows = []
for model_name, path in MODELS.items():
    if path.exists():
        df = pd.read_csv(path, sep=';')
        df.insert(0, "Model", model_name)
        rows.append(df)
        print(f"✓ {model_name}: {path}")
    else:
        print(f"✗ {model_name}: {path} nicht gefunden, wird übersprungen")

if not rows:
    print("Keine Dateien gefunden. Abbruch.")
else:
    OUTPUT_FILE = get_output_path()
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    df_combined = pd.concat(rows, ignore_index=True)
    df_combined.to_excel(OUTPUT_FILE, index=False)

    from openpyxl import load_workbook
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(OUTPUT_FILE)
    print(f"\n✓ Gespeichert: {OUTPUT_FILE.resolve()}")